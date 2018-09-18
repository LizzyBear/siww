'''
将company database 文件中的公司按照设定关键词分类统计并生成表格
'''

import re
from openpyxl import load_workbook

#打开company database.xlsx
wb = load_workbook(filename = 'Company database_water industry.xlsx')
sheet = wb.get_sheet_by_name('crawler output')
rn = sheet.max_row


#建立一个新的sheet用于放关键词筛选后的数据
wb.create_sheet('ns1')
ns1 = wb.get_sheet_by_name('ns1')

#将column E中的公司简介提取为一个list
def get_company_info():
    company_info = []
    for r in range (1, rn+1):
        info = sheet.cell(row = r, column = 5).value
        company_info.append(info)
    return company_info



#对sheet中的公司简介column进行关键词搜索并形成结果
# def keyword_sorting():
#     keyword_template = re.findall('[\s\S]*(membrane)?[\s\S]*', 











